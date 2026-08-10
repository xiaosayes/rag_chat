# -*- coding: utf-8 -*-
"""
全面代码审查（2026-08-10）新增缺陷的固化测试 —— 测试工程师视角

约定：
  - 每个测试断言【期望的正确行为】；
  - 当前未修复时测试 FAIL（即复现缺陷），修复后转 PASS；
  - 测试不依赖任何外部 API（全部 mock / 离线构造）。

覆盖缺陷清单（详见审查报告）：
  F1  xlsx 首行全空 → 整个 sheet 数据静默丢失
  F2  ASR 容器魔数检测遗漏 MP4/M4A（ftyp 在偏移 4，非偏移 0）→ Safari 录音被当裸 PCM
  F3  ASR 响应 ws[].cw 为空 → IndexError，接收线程静默死亡
  F4  混合检索把"瞬时故障的不完整结果"写入缓存（5 分钟内故障固化）
  F5  rerank 本地降级遇全空文本 → TF-IDF 空词表 ValueError，穿透降级保护
  F6  rerank 单候选提前返回 RRF 分，pipeline 标记 reranked=True → 相关性闸门误判
  F7  VectorStore.search 单条 payload 损坏（metadata_json=null/坏串）→ 整个语义检索崩溃
  F8  VectorStore.close() 后 client 属性静默返回 None → 下游 AttributeError
  F9  embedding batch_size <= 0 未防御 → embed_batch 崩溃
  F11 format_answer 把 <0.1 的重排低分误判为 RRF 量纲 → 相关度标签错误
  F12 _select_prompt 缺 CHITCHAT 键 → 潜在 KeyError
  F13 Artifact.to_text 遇非字符串 tags → TypeError（latent，函数当前未被调用）
  F14 clean_text_for_tts：货币$在前、LaTeX$在后 → 公式未被转换、残留裸 $
  F15 L0 is_kb_related 漏判"谢谢你/感谢你的帮助"（语义层兜底关闭时走完整 RAG）
"""
import sys
import tempfile
from pathlib import Path
from unittest.mock import MagicMock

import pytest

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from src.chunking import Chunk
from src.cache import retrieval_cache


# =============================================================================
# F1: xlsx 首行全空 → 整表数据静默丢失
# =============================================================================
class TestXlsxBlankFirstRow:
    """真实 Excel 常见场景：首行为空行/标题行，数据表头在第 2+ 行"""

    def _make_xlsx(self, rows, tmp_path):
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        for r in rows:
            ws.append(r)
        p = tmp_path / "t.xlsx"
        wb.save(str(p))
        return p

    def test_blank_first_row_should_not_lose_data(self, tmp_path):
        """首行全空时，空行应被跳过，第 2 行表头生效；当前整表 0 条记录（数据丢失）"""
        from src.data_loader import DataLoader
        p = self._make_xlsx([
            [None, None],            # 空行（真实表格常见）
            ["名称", "描述"],         # 真正的表头
            ["司母戊鼎", "商代青铜器"],
        ], tmp_path)
        arts = DataLoader.load(p)
        names = [a.name for a in arts]
        assert "司母戊鼎" in names, f"整表数据被静默丢弃: {names}"

    def test_normal_header_still_works(self, tmp_path):
        """回归保护：首行即表头的正常文件不受影响"""
        from src.data_loader import DataLoader
        p = self._make_xlsx([
            ["名称", "描述"],
            ["司母戊鼎", "商代青铜器"],
        ], tmp_path)
        arts = DataLoader.load(p)
        assert [a.name for a in arts] == ["司母戊鼎"]


# =============================================================================
# F2: ASR 容器魔数检测遗漏 MP4/M4A（ftyp 位于偏移 4）
# =============================================================================
class TestAsrContainerMagic:
    def test_mp4_ftyp_at_offset4_detected(self):
        """ISO-BMFF（mp4/m4a）前 4 字节是 box size，'ftyp' 在偏移 4；
        当前只检查偏移 0 → Safari 录音被当裸 PCM 送讯飞，识别输出乱码"""
        from src.asr import _is_encoded_container
        mp4 = b"\x00\x00\x00\x18ftypmp42\x00\x00\x00\x00mp42isom" + b"\x00" * 64
        assert _is_encoded_container(mp4), "MP4/M4A 容器未识别（ftyp 在偏移 4）"

    def test_webm_ogg_mp3_still_detected(self):
        """回归保护：已有魔数不受影响"""
        from src.asr import _is_encoded_container
        assert _is_encoded_container(b"\x1aE\xdf\xa3" + b"\x00" * 16)   # webm
        assert _is_encoded_container(b"OggS" + b"\x00" * 16)           # ogg
        assert _is_encoded_container(b"ID3" + b"\x00" * 16)            # mp3
        assert not _is_encoded_container(b"RIFF" + b"\x00" * 16)       # wav 不算编码容器


# =============================================================================
# F3: ASR 响应 ws[].cw 为空 → IndexError（接收线程静默死亡，识别卡死到超时）
# =============================================================================
class TestAsrMalformedFrame:
    def test_empty_cw_does_not_crash(self):
        from src.asr import IflytekASR
        asr = IflytekASR("app", "key", "secret", _ws=None)
        # 不应抛异常（当前 IndexError: list index out of range）
        asr._handle_message('{"code":0,"data":{"result":{"ws":[{"cw":[]}]}}}')

    def test_missing_cw_key_does_not_crash(self):
        from src.asr import IflytekASR
        asr = IflytekASR("app", "key", "secret", _ws=None)
        asr._handle_message('{"code":0,"data":{"result":{"ws":[{}]}}}')

    def test_normal_frame_still_works(self):
        from src.asr import IflytekASR
        asr = IflytekASR("app", "key", "secret", _ws=None)
        asr._handle_message(
            '{"code":0,"data":{"result":{"ws":[{"cw":[{"w":"你好"}]}],"ls":true}}}'
        )
        assert asr.final_text == "你好"
        assert asr.is_final()


# =============================================================================
# F4: 混合检索缓存"故障结果"（瞬时 API 故障被固化 5 分钟）
# =============================================================================
class TestRetrieveDoesNotCacheFailures:
    def _make_retriever(self):
        from src.retriever import HybridRetriever
        vs = MagicMock(); vs.collection_name = "audit_test"
        return HybridRetriever(vs, MagicMock(), MagicMock())

    def test_partial_failure_not_cached(self):
        """语义检索失败（BM25 成功）的结果不应写入缓存；
        当前故障结果被缓存，语义恢复后 5 分钟内仍只返回 BM25 结果"""
        retrieval_cache.clear()
        hr = self._make_retriever()
        sem_chunk = Chunk(id="s1", artifact_id="a", artifact_name="s", text="sem")
        bm_chunk = Chunk(id="b1", artifact_id="a", artifact_name="b", text="bm")
        hr._semantic_search = MagicMock(
            side_effect=[RuntimeError("API down"), [(sem_chunk, 0.9)]]
        )
        hr._bm25_search = MagicMock(return_value=[(bm_chunk, 0.5)])
        r1 = hr.retrieve("审计测试问题A")
        assert [c.id for c, _ in r1] == ["b1"]          # 第一次确实故障降级
        r2 = hr.retrieve("审计测试问题A")                 # 语义已恢复
        ids2 = [c.id for c, _ in r2]
        assert "s1" in ids2, f"故障结果被缓存，恢复后仍返回旧结果: {ids2}"

    def test_total_failure_not_cached(self):
        """双侧检索都失败返回的空结果不应写入缓存（否则 TTL 内永远空）"""
        retrieval_cache.clear()
        hr = self._make_retriever()
        bm_chunk = Chunk(id="b9", artifact_id="a", artifact_name="b", text="bm")
        # 每次 retrieve 各调用一次：第 1 次双侧故障，第 2 次双侧恢复
        hr._semantic_search = MagicMock(side_effect=[RuntimeError("down"), []])
        hr._bm25_search = MagicMock(
            side_effect=[RuntimeError("down"), [(bm_chunk, 0.5)]]
        )
        assert hr.retrieve("审计测试问题B") == []
        r2 = hr.retrieve("审计测试问题B")
        assert [c.id for c, _ in r2] == ["b9"], "空故障结果被缓存，恢复后仍为空"


# =============================================================================
# F5: rerank 本地降级遇全空文本 → TF-IDF 空词表崩溃，穿透 rerank() 降级保护
# =============================================================================
class TestRerankLocalEmptyVocab:
    def test_empty_texts_do_not_crash_local_rerank(self):
        from src.reranker import BailianReranker
        r = BailianReranker(api_key="fake")
        cands = [
            (Chunk(id="1", artifact_id="a", artifact_name="a", text=""), 0.1),
            (Chunk(id="2", artifact_id="b", artifact_name="b", text=""), 0.2),
        ]
        # 当前 ValueError: empty vocabulary；期望降级为原顺序返回
        out = r._rerank_local("", cands)
        assert len(out) == 2

    def test_rerank_full_path_survives_api_down_and_empty_texts(self):
        """API 失败 + 空文本：rerank() 应原样降级返回，不向 pipeline 抛异常"""
        from src.reranker import BailianReranker
        r = BailianReranker(api_key="fake")
        r._rerank_with_api = MagicMock(side_effect=RuntimeError("API down"))
        cands = [(Chunk(id="1", artifact_id="a", artifact_name="a", text=""), 0.1)]
        out = r.rerank("问题", cands)   # 当前 len<=1 提前返回，不触发；2 条时触发
        cands2 = cands + [(Chunk(id="2", artifact_id="b", artifact_name="b", text=""), 0.2)]
        out = r.rerank("问题", cands2)
        assert len(out) == 2


# =============================================================================
# F6: rerank 单候选提前返回 → 分数仍为 RRF 量级；pipeline 按 0~1 阈值误判
# =============================================================================
class TestRerankSingleCandidateScoreScale:
    def test_single_candidate_early_return_score_semantics(self):
        """rerank() 对单候选不应直接返回原始（RRF）分数；
        pipeline 的 _has_relevant_results 将其与 RELEVANCE_THRESHOLD=0.45 比较会误判。
        修复后：单候选也走 API 拿真实 0~1 相关性分。"""
        from src.reranker import BailianReranker
        r = BailianReranker(api_key="fake")
        cands = [(Chunk(id="1", artifact_id="a", artifact_name="a", text="相关"), 0.008)]
        # 模拟 API 返回真实相关性分（0~1）
        r._rerank_with_api = MagicMock(return_value=[(cands[0][0], 0.9)])
        out = r.rerank("问题", cands)
        r._rerank_with_api.assert_called_once()  # 单候选不再提前返回
        assert out[0][1] > 0.1, (
            f"单候选返回 RRF 分 {out[0][1]}，pipeline 将按 0.45 阈值误判为不相关"
        )

    def test_single_candidate_api_down_falls_back_gracefully(self):
        """单候选 + API 故障 → 本地降级/原序返回，不抛异常"""
        from src.reranker import BailianReranker
        r = BailianReranker(api_key="fake")
        r._rerank_with_api = MagicMock(side_effect=RuntimeError("API down"))
        cands = [(Chunk(id="1", artifact_id="a", artifact_name="a", text="相关文档"), 0.008)]
        out = r.rerank("问题", cands)
        assert len(out) == 1 and out[0][0].id == "1"


# =============================================================================
# F7: VectorStore.search 单条 payload 损坏 → 整个语义检索崩溃
# =============================================================================
class TestVectorStoreSearchRobustness:
    def _vs_with_hits(self, payloads):
        from src.vector_store import VectorStore
        vs = VectorStore.__new__(VectorStore)
        vs.collection_name = "audit"
        vs.local_mode = True
        hits = []
        for i, pl in enumerate(payloads):
            h = MagicMock()
            h.payload = pl
            h.score = 0.9 - i * 0.1
            hits.append(h)
        resp = MagicMock(); resp.points = hits
        client = MagicMock()
        client.query_points.return_value = resp
        vs._client = client
        vs._closed = False
        return vs

    def test_null_metadata_json_skips_bad_point(self):
        """metadata_json=null（JSON null）：核心字段完好，容忍为空 metadata 保留该点；
        修复前 json.loads(None) 抛 TypeError 使整条 search 失败（语义检索静默为空）"""
        good = {"chunk_id": "c1", "text": "t", "metadata_json": "{}"}
        null_meta = {"chunk_id": "c2", "text": "t", "metadata_json": None}
        vs = self._vs_with_hits([null_meta, good])
        results = vs.search([0.1] * 8, top_k=2)
        assert [c.id for c, _ in results] == ["c2", "c1"]  # 均不崩溃，null→{}
        assert results[0][0].metadata == {}

    def test_corrupt_metadata_json_string_skips_bad_point(self):
        good = {"chunk_id": "c1", "text": "t", "metadata_json": "{}"}
        bad = {"chunk_id": "c2", "text": "t", "metadata_json": "{not-json"}
        vs = self._vs_with_hits([bad, good])
        results = vs.search([0.1] * 8, top_k=2)
        assert [c.id for c, _ in results] == ["c1"]


# =============================================================================
# F8: VectorStore.close() 后 client 静默返回 None → 下游 AttributeError
# =============================================================================
class TestVectorStoreClosedClient:
    def test_client_after_close_raises_clear_error(self):
        """关闭后访问 client 应抛出语义清晰的 RuntimeError，而不是静默返回 None
        （None 透传到 search/upsert 会报 'NoneType' object has no attribute ...）"""
        import threading
        from src.vector_store import VectorStore
        vs = VectorStore.__new__(VectorStore)
        vs._client = None
        vs._closed = True
        vs._connect_lock = threading.Lock()
        with pytest.raises(RuntimeError, match="已关闭"):
            _ = vs.client

    def test_reset_connection_allows_reconnect(self):
        """回归保护：reset_connection 后 _closed 复位（锁内完成），可重新连接"""
        import threading
        from src.vector_store import VectorStore
        vs = VectorStore.__new__(VectorStore)
        vs._client = None  # close() 在 _client=None 时无副作用
        vs._closed = True
        vs._connect_lock = threading.Lock()
        VectorStore.reset_connection(vs)
        assert vs._closed is False


# =============================================================================
# F9: embedding batch_size <= 0 未防御
# =============================================================================
class TestEmbeddingBatchSizeNonPositive:
    def test_zero_batch_size_clamped(self):
        from src.embeddings import BailianEmbedding
        e = BailianEmbedding(api_key="fake", batch_size=0)
        assert e.batch_size >= 1, "batch_size=0 未钳制，embed_batch 内 range(step=0) 崩溃"

    def test_negative_batch_size_clamped(self):
        from src.embeddings import BailianEmbedding
        e = BailianEmbedding(api_key="fake", batch_size=-3)
        assert e.batch_size >= 1, "batch_size<0 未钳制，embed_batch 全部批缺失报 RuntimeError"


# =============================================================================
# F11: format_answer 把 <0.1 的重排低分误判为 RRF 量纲 → 第 1 名标 [高]
# =============================================================================
class TestFormatAnswerScoreScale:
    def test_low_rerank_score_labeled_low(self):
        import app
        chunks = [{"artifact_name": "A", "score": 0.05, "chunk_type": "detail"}]
        out = app.format_answer("回答", chunks)
        # 0.05 是重排量纲（0~1）的低相关分，应标 [低]；
        # 当前 0<0.05<0.1 被当成 RRF 量纲 → 按排名标 [高]
        assert "[高]" not in out, f"重排低分被误标为[高]: {out}"


# =============================================================================
# F12: _select_prompt 缺 CHITCHAT 键 → 潜在 KeyError
# =============================================================================
class TestSelectPromptChitchatKey:
    def test_chitchat_type_has_prompt_mapping(self):
        from src.rag_pipeline import RAGPipeline, QueryType
        p = RAGPipeline.__new__(RAGPipeline)
        p.project_cfg = None
        # 当前 KeyError；期望回退到 default prompt
        prompt = p._select_prompt(QueryType.CHITCHAT, "上下文")
        assert "上下文" in prompt


# =============================================================================
# F13: Artifact.to_text 遇非字符串 tags → TypeError（latent）
# =============================================================================
class TestArtifactToTextNonStrTags:
    def test_numeric_tags_do_not_crash(self):
        from src.data_loader import Artifact
        a = Artifact(name="司母戊鼎", tags=[1, 2, 3])  # JSON 数字标签
        text = a.to_text()  # 当前 TypeError
        assert "司母戊鼎" in text


# =============================================================================
# F14: clean_text_for_tts 货币$在前 LaTeX$在后 → 公式残留
# =============================================================================
class TestTtsCleanDollarLatex:
    def test_currency_before_latex(self):
        from src.utils import clean_text_for_tts
        out = clean_text_for_tts("价格 $5 元，公式 $x^2$ 结束")
        assert "x^2" not in out, f"LaTeX 公式残留未转换: {out!r}"
        assert "$5" in out  # 货币保留

    def test_latex_alone_still_converted(self):
        """回归保护：单独的 LaTeX 正常转换"""
        from src.utils import clean_text_for_tts
        out = clean_text_for_tts("公式 $x^2$ 结束")
        assert "x 的平方" in out


# =============================================================================
# F15: L0 is_kb_related 漏判"谢谢你"类闲聊
# =============================================================================
class TestIsKbRelatedThanks:
    @pytest.mark.parametrize("q", ["谢谢你", "感谢你", "谢谢您", "多谢啦"])
    def test_thanks_not_kb_related(self, q):
        """L0 规则应识别纯感谢为闲聊；此前漏判（依赖 L1 语义兜底，
        但 INTENT_SEMANTIC_ENABLED=false 时会走完整 RAG 检索+LLM）。
        注："感谢你的帮助" 剥离后剩 "你的帮助" 有实质内容，判知识库相关是
        前一轮审查固化（test_review_findings.TestIsKBRelatedFalsePositives）的
        可接受行为，不在本修复范围内。"""
        from src.rag_pipeline import RAGPipeline
        assert RAGPipeline.is_kb_related(q) is False

    def test_thanks_with_substance_still_kb_related(self):
        """回归保护（前轮审查契约）：含实质内容的感谢句仍路由知识库"""
        from src.rag_pipeline import RAGPipeline
        assert RAGPipeline.is_kb_related("谢谢你的帮助") is True
        assert RAGPipeline.is_kb_related("谢谢你的帮助是什么文物") is True


# =============================================================================
# 修复后新增回归测试（F12/F19/F20/F21/F23/F24/F25）
# =============================================================================
class TestTtsReplayUniqueFile:
    """F12：TTS 重播文件按请求唯一命名，多用户不互相覆盖"""

    def _make_wav(self):
        import io
        import wave
        buf = io.BytesIO()
        with wave.open(buf, "wb") as w:
            w.setnchannels(1)
            w.setsampwidth(2)
            w.setframerate(24000)
            w.writeframes(b"\x00\x00" * 2400)
        return buf.getvalue()

    def test_replay_files_unique(self, tmp_path, monkeypatch):
        from src.config import settings
        monkeypatch.setattr(settings, "project_root", tmp_path)
        import app
        wav = self._make_wav()
        p1 = app._write_replay_wav([wav])
        p2 = app._write_replay_wav([wav])
        assert p1 != p2, "重播文件应唯一命名，避免多用户互相覆盖"
        assert p1.exists() and p2.exists()


class TestPptFallback:
    """F20：.ppt 旧格式友好降级（与 DocxParser 的 .doc 处理一致）"""

    def test_ppt_does_not_raise(self, tmp_path):
        from src.document_loader import PptxParser
        p = tmp_path / "old.ppt"
        p.write_bytes(b"\xd0\xcf\x11\xe0 not-a-real-ppt")
        doc = PptxParser().parse(p)  # 不再抛异常
        assert doc.content


class TestProjectIdValidation:
    """F21：外部项目 JSON 的非法 id 被拒绝（路径穿越防护与 add_project 一致）"""

    def test_malicious_id_rejected(self, tmp_path):
        import json
        from src.project import ProjectManager
        d = tmp_path / "projects"
        d.mkdir()
        (d / "evil.json").write_text(
            json.dumps({"id": "../../evil", "name": "evil"}), encoding="utf-8"
        )
        pm = ProjectManager(projects_dir=d)
        assert "../../evil" not in pm._projects


class TestLLMIntentNegation:
    """F23：LLM 意图分类否定表述不误命中子串"""

    def test_not_chitchat_returns_none(self):
        from src.intent_classifier import classify_with_llm
        llm = MagicMock()
        llm.chat.return_value = "not chitchat"
        import src.intent_classifier as ic
        # 绕过 api_key 检查
        from src.config import settings
        if not settings.dashscope_api_key:
            pytest.skip("需要 DASHSCOPE_API_KEY 配置")
        assert classify_with_llm(llm, "司母戊鼎有多重") is None

    def test_exact_intent_still_works(self):
        from src.intent_classifier import classify_with_llm
        llm = MagicMock()
        llm.chat.return_value = "factual"
        from src.config import settings
        if not settings.dashscope_api_key:
            pytest.skip("需要 DASHSCOPE_API_KEY 配置")
        assert classify_with_llm(llm, "司母戊鼎有多重") == "factual"


class TestGreetingBoundary:
    """F24：问候词边界匹配，子串不再误伤"""

    def test_cunzai_ma_not_greeting(self):
        from src.rag_pipeline import RAGPipeline, QueryType
        p = RAGPipeline.__new__(RAGPipeline)
        assert p._is_greeting("这件文物现在还存在吗") is False  # "存在吗"含"在吗"
        assert p._is_greeting("在吗") is True
        assert p._is_greeting("在吗？") is True
        assert p._is_greeting("你好，请问") is True


class TestCacheAtomicSave:
    """F25：缓存原子写（tmp + replace），不留半截文件"""

    def test_save_produces_valid_json(self):
        import json
        from src.cache import EmbeddingCache
        d = Path(tempfile.mkdtemp())
        cache = EmbeddingCache(cache_dir=d)
        cache.set("问题A", [0.1, 0.2])
        cache.save()
        data = json.loads((d / "exact_cache.json").read_text(encoding="utf-8"))
        assert data["问题A"] == [0.1, 0.2]
        assert not list(d.glob("*.tmp")), "临时文件应已被原子替换"


if __name__ == "__main__":
    pytest.main([__file__, "-v", "--tb=short"])
