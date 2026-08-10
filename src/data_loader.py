"""
数据加载与预处理模块
支持从 JSON / CSV 等格式加载文物数据，并进行标准化处理
"""

from __future__ import annotations

import json
import csv
import uuid
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List, Optional
from dataclasses import dataclass, field, asdict

from loguru import logger

from src.utils import generate_id, load_json


# ========== Excel (.xlsx) 支持（bug-109） ==========
# 名称列候选：命中即作为记录名称（无法穷举所有项目列名，
# 未命中的列会全部拼入 description 参与全文检索）
_NAME_COLUMN_CANDIDATES = {
    "名称", "name", "Name", "标题", "title", "Title",
    "展商名称", "企业名称", "公司名称", "项目名称",
    "品牌名称", "物料名称", "产品名称", "文件名称", "文档名称",
}


def _cell_to_str(value: Any) -> str:
    """Excel 单元格值统一转字符串（数字/布尔/日期）"""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "是" if value else "否"
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


def _find_name_column(record: Dict[str, Any]) -> Optional[str]:
    """识别记录中的名称列（候选列名优先，None 表示未识别）"""
    for cand in _NAME_COLUMN_CANDIDATES:
        if cand in record:
            return cand
    return None


@dataclass
class Artifact:
    """文物数据模型"""
    id: str = ""
    name: str = ""                          # 名称
    alias: str = ""                         # 别名
    dynasty: str = ""                       # 朝代
    category: str = ""                      # 类别（青铜器/瓷器/书画/玉器等）
    material: str = ""                      # 材质
    year: str = ""                          # 年代/年份
    provenance: str = ""                    # 出土地点
    location: str = ""                      # 现藏地点
    description: str = ""                   # 描述
    historical_significance: str = ""       # 历史意义
    cultural_value: str = ""                # 文化价值
    tags: List[str] = field(default_factory=list)      # 标签
    importance: int = 3                     # 重要性 1-5
    image_url: str = ""                     # 图片 URL
    source: str = ""                        # 数据来源
    extra: Dict[str, Any] = field(default_factory=dict)  # 扩展字段

    def __post_init__(self):
        if not self.id:
            # 使用多字段组合生成 ID，降低同名同朝代文物碰撞概率
            base = self.name + self.dynasty + self.category + self.material
            if base:
                self.id = generate_id(base)
            else:
                # bug-061 修复：全空字段时追加随机串，避免多件空文物共享 md5("") 导致 ID 碰撞、向量互相覆盖
                self.id = generate_id(base + uuid.uuid4().hex)

    def to_dict(self) -> Dict[str, Any]:
        return asdict(self)

    def to_text(self) -> str:
        """转换为文本表示（用于检索）"""
        parts = [
            f"名称：{self.name}",
            f"朝代：{self.dynasty}",
            f"类别：{self.category}",
            f"材质：{self.material}",
            f"现藏：{self.location}",
            f"描述：{self.description}",
        ]
        if self.historical_significance:
            parts.append(f"历史意义：{self.historical_significance}")
        if self.cultural_value:
            parts.append(f"文化价值：{self.cultural_value}")
        if self.tags:
            # audit-F13 修复：tags 可能含非字符串元素（JSON 数字列表），统一转 str，
            # 避免 join 抛 TypeError（与 chunking.py bug-090 修复一致）
            parts.append(f"标签：{'、'.join(str(t) for t in self.tags)}")
        return "\n".join(parts)


class DataLoader:
    """数据加载器"""

    SUPPORTED_FORMATS = {"json", "csv", "xlsx"}

    @staticmethod
    def load(path: Path, format: Optional[str] = None) -> List[Artifact]:
        """加载数据文件"""
        if format is None:
            format = path.suffix.lstrip(".").lower()

        if format not in DataLoader.SUPPORTED_FORMATS:
            raise ValueError(f"不支持的文件格式: {format}，支持: {DataLoader.SUPPORTED_FORMATS}")

        loader_map = {
            "json": DataLoader._load_json,
            "csv": DataLoader._load_csv,
            "xlsx": DataLoader._load_xlsx,
        }

        loader = loader_map[format]
        raw_data = loader(path)
        artifacts = [DataLoader._normalize(item) for item in raw_data]
        logger.info(f"成功加载 {len(artifacts)} 件文物数据")
        return artifacts

    @staticmethod
    def _load_json(path: Path) -> List[Dict[str, Any]]:
        """加载 JSON 格式"""
        return load_json(path)

    @staticmethod
    def _load_csv(path: Path) -> List[Dict[str, Any]]:
        """加载 CSV 格式"""
        if not path.exists():
            raise FileNotFoundError(f"文件不存在: {path}")
        with open(path, "r", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            data = list(reader)
        logger.info(f"已加载 {len(data)} 条数据 from {path}")
        return data

    @staticmethod
    def _load_xlsx(path: Path) -> List[Dict[str, Any]]:
        """
        加载 Excel (.xlsx) 格式（bug-109）

        规则：
          - 遍历所有 sheet（多 sheet 支持），每个 sheet 第一个非空行为表头，后续行每行一条记录
          - 名称列识别：列名命中候选集 → 否则取第一个非空列
          - 任意未识别列以 "列名：值" 拼入 description，保证任何列内容可被全文检索命中
          - 空行 / 空列跳过
        """
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise ImportError(
                "请先安装 openpyxl: pip install openpyxl（Excel 数据源支持）"
            )
        if not path.exists():
            raise FileNotFoundError(f"文件不存在: {path}")

        records: List[Dict[str, Any]] = []
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            for ws in wb.worksheets:
                rows = ws.iter_rows(values_only=True)
                header: Optional[List[str]] = None
                for row in rows:
                    if header is None:
                        # audit-F1 修复：首行全空时 ["","",...] 为非空列表，旧判断
                        # `if not header` 会把它误认为有效表头 → 真正的表头行被当数据、
                        # 列名全空被跳过 → 整个 sheet 静默丢失 0 条记录。
                        # 改为 None 标志位 + 跳过前导空行（真实 Excel 常见标题/空行）。
                        candidate = [_cell_to_str(c) for c in row]
                        if not any(candidate):
                            continue
                        header = candidate
                        continue
                    # 空行跳过
                    if all(_cell_to_str(c) == "" for c in row):
                        continue
                    record: Dict[str, Any] = {}
                    for col_name, value in zip(header, row):
                        col_name = col_name.strip()
                        if not col_name:
                            continue
                        text = _cell_to_str(value)
                        if text:
                            record[col_name] = text
                    if not record:
                        continue
                    # 名称列识别：候选列名 → 第一非空列 → 兜底
                    name_col = _find_name_column(record)
                    if name_col:
                        record["name"] = record.pop(name_col)
                    else:
                        first_key = next(iter(record))
                        record["name"] = record.pop(first_key)
                    # 描述拼接：所有非 name 列以 "列名：值" 拼入，保证任意列可检索
                    parts = [
                        f"{k}：{v}" for k, v in record.items() if k != "name" and v
                    ]
                    record["description"] = "\n".join(parts)
                    record["sheet"] = ws.title
                    records.append(record)
        finally:
            wb.close()
        logger.info(f"已加载 {len(records)} 条数据 from {path}")
        return records

    @staticmethod
    def _normalize(item: Dict[str, Any]) -> Artifact:
        """标准化单条文物数据"""
        # 字段映射（兼容不同命名）
        field_map = {
            "name": ["name", "名称", "文物名称", "文物名", "Name"],
            "alias": ["alias", "别名", "又称", "亦称"],
            "dynasty": ["dynasty", "朝代", "年代", "时期", "时代"],
            "category": ["category", "类别", "分类", "类型", "种类"],
            "material": ["material", "材质", "材料", "质地"],
            "year": ["year", "年份", "具体年代", "公元"],
            "provenance": ["provenance", "出土地", "出土地点", "出土"],
            "location": ["location", "现藏", "收藏", "藏于", "博物馆", "所在地"],
            "description": ["description", "描述", "简介", "介绍", "说明", "概述"],
            "historical_significance": ["historical_significance", "历史意义", "历史价值", "意义"],
            "cultural_value": ["cultural_value", "文化价值", "艺术价值", "科学价值"],
            "tags": ["tags", "标签", "关键词", "关键字"],
            "importance": ["importance", "重要性", "重要程度", "等级", "级别"],
            "image_url": ["image_url", "图片", "图片URL", "image"],
            "source": ["source", "来源", "数据来源"],
        }

        normalized = {}
        for target_key, source_keys in field_map.items():
            for sk in source_keys:
                if sk in item and item[sk]:
                    normalized[target_key] = item[sk]
                    break

        # 处理 tags 字段（可能是字符串列表或逗号分隔）
        if "tags" in normalized:
            if isinstance(normalized["tags"], str):
                normalized["tags"] = [
                    t.strip() for t in normalized["tags"].split(",") if t.strip()
                ]

        # 处理 importance
        if "importance" in normalized:
            try:
                normalized["importance"] = int(float(normalized["importance"]))
            except (ValueError, TypeError):
                normalized["importance"] = 3
        else:
            normalized["importance"] = 3

        # 收集 extra 字段
        used_keys = set()
        for keys in field_map.values():
            used_keys.update(keys)
        extra = {k: v for k, v in item.items() if k not in used_keys and v}

        return Artifact(**normalized, extra=extra)